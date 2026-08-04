#!/usr/bin/env python3
"""
publish.py -- turn every spreadsheet in inbox/ into a folder of maps under site/.

    python3 pipeline/publish.py                 # everything in inbox/
    python3 pipeline/publish.py file.xlsx       # one file
    python3 pipeline/publish.py --tabs 534_ORLANDO

Each tab becomes site/<slug>/index.html, so a market keeps the same URL from one
drop to the next and the Google Sites embeds never need touching.

A tab whose column layout nothing recognizes is handed to resolve_schema.py, which
learns it into schemas.json. That file is committed, so the same layout is only
ever resolved once. If it can't be resolved the tab is reported and skipped -- the
run continues with the tabs that did work rather than failing whole.

Publishing is not done here. In CI the workflow uploads site/ to Pages; run
validate.py first, because it is what decides whether any of this is fit to ship.
"""
import argparse
import glob
import os
import re
import shutil
import subprocess
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
VERBOSE = False
SKIP_TABS = {"splitcode", "readme", "notes", "cover", "instructions"}

sys.path.insert(0, HERE)
from clean_addresses import build_mapping  # noqa: E402


def sh(cmd, cwd=None, check=True, quiet=False):
    r = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True)
    if not quiet and r.stdout:
        print("    " + r.stdout.strip().replace("\n", "\n    "))
    if check and r.returncode != 0:
        print((r.stderr or r.stdout).strip(), file=sys.stderr)
        raise SystemExit(f"failed: {' '.join(str(c) for c in cmd)}")
    return r


# ---------------------------------------------------------------- product naming
# Filenames follow "<date>_<PRODUCT>_Maps & Addresses.xlsx", so the product falls
# out of the name and never has to be typed. Anything a product folder cannot be
# derived from lands in "unsorted", which is visible rather than silent.
DEDUP_SUFFIX = re.compile(r"\s*\(\d+\)\s*$")
RESPONDENT = re.compile(r"\s+-\s+[A-Za-z]+(?:\s+[A-Za-z]+){0,3}\s*$")
DATE_PREFIX = re.compile(r"^(?:\d{4}[-_]\d{1,2}|\d{1,2}[-_]\d{2,4})[-_\s]+")
BOILERPLATE = re.compile(r"[_\s-]*maps?\s*&\s*addresses.*$", re.I)


def product_from_filename(path):
    """8_26_NG OOF_Maps & Addresses (1).xlsx -> ng-oof"""
    stem = os.path.splitext(os.path.basename(path))[0]
    for rx in (DEDUP_SUFFIX, RESPONDENT, DEDUP_SUFFIX):
        stem = rx.sub("", stem)
    stem = DATE_PREFIX.sub("", stem)
    stem = BOILERPLATE.sub("", stem)
    return slugify(stem) or "unsorted"


def discover(paths):
    """[(product, path)] for every spreadsheet given, expanding folders.

    A file inside inbox/<product>/ takes that folder as its product -- the folder
    is the authority, so a misfiled upload is fixed by moving it. A file sitting
    at the top of inbox/ has its product inferred from its name instead.
    """
    out = []
    for i in paths:
        if os.path.isdir(i):
            for entry in sorted(os.listdir(i)):
                sub = os.path.join(i, entry)
                if os.path.isdir(sub):
                    for f in sorted(os.listdir(sub)):
                        fp = os.path.join(sub, f)
                        if os.path.isfile(fp) and is_sheet(fp):
                            out.append((slugify(entry), fp))
                elif os.path.isfile(sub) and is_sheet(sub):
                    out.append((product_from_filename(sub), sub))
        elif os.path.exists(i):
            out.append((product_from_filename(i), i))
        else:
            raise SystemExit(f"no such file or folder: {i}")
    return [(p, f) for p, f in out
            if not os.path.basename(f).startswith("~$")]


def is_sheet(path):
    return os.path.splitext(path)[1].lower() in (".xlsx", ".xlsm", ".xls", ".csv",
                                                 ".tsv", ".txt")


def slugify(tab):
    """534_ORLANDO -> 534-orlando. Keeps the DMA code so the URL is unambiguous
    and matches the tab name in the source file."""
    return re.sub(r"-+", "-", re.sub(r"[^A-Za-z0-9]+", "-", str(tab)).strip("-")).lower()


def product_label(product):
    """ng-oof -> NG OOF"""
    return " ".join(w.upper() for w in str(product).split("-") if w)


def titleize(tab, product=""):
    """534_ORLANDO -> Orlando (534) -- NG OOF

    The product belongs in the title now: the same city appears under several
    product lines, and a map with no product on it is indistinguishable from
    another product's map of the same city.
    """
    m = re.match(r"^(\d+)[_\s-]+(.*)$", str(tab).strip())
    code, rest = (m.group(1), m.group(2)) if m else ("", str(tab))
    words = " ".join(w.capitalize() for w in re.split(r"[_\s]+", rest) if w)
    base = f"{words} ({code})" if code else words
    lab = product_label(product)
    return f"{base} \u2014 {lab}" if lab and lab != "UNSORTED" else base


DELIMS = {".tsv": "\t", ".csv": ",", ".txt": ","}


def is_flat(path):
    """A single-table file (csv/tsv/txt) rather than a workbook of tabs."""
    return os.path.splitext(path)[1].lower() in DELIMS


def read_headers_flat(path):
    """Headers, row count and modal state for a csv/tsv. There are no tabs, so
    the market name comes from the filename instead."""
    import csv as _csv
    delim = DELIMS[os.path.splitext(path)[1].lower()]
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        rd = _csv.reader(f, delimiter=delim)
        headers = []
        for row in rd:
            if row and sum(1 for c in row if str(c).strip()) >= 2:
                headers = [str(c or "").strip() for c in row]
                break
        if not headers:
            return [], 0, ""
        norm = [re.sub(r"[^A-Z0-9]", "", h.upper()) for h in headers]
        si = next((i for i, x in enumerate(norm)
                   if x in ("TSTATE", "STATE", "SERVICESTATE", "SERVICESTATECD")), None)
        n, states = 0, {}
        for row in rd:
            if not row or all(not str(c).strip() for c in row):
                continue
            n += 1
            if si is not None and si < len(row) and row[si]:
                v = str(row[si]).strip().upper()[:2]
                states[v] = states.get(v, 0) + 1
    return headers, n, (max(states, key=states.get) if states else "")


def read_headers(xlsx, tab):
    if is_flat(xlsx):
        return read_headers_flat(xlsx)
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[tab]
    headers, n, states = None, 0, {}
    it = ws.iter_rows(values_only=True)
    for row in it:
        if row and sum(1 for c in row if c not in (None, "")) >= 2:
            headers = [str(c or "") for c in row]
            break
    if headers:
        norm = [re.sub(r"[^A-Z0-9]", "", h.upper()) for h in headers]
        si = next((i for i, x in enumerate(norm)
                   if x in ("TSTATE", "STATE", "SERVICESTATE", "SERVICESTATECD")), None)
        for row in it:
            if row is None or all(c in (None, "") for c in row):
                continue
            n += 1
            if si is not None and si < len(row) and row[si]:
                v = str(row[si]).strip().upper()[:2]
                states[v] = states.get(v, 0) + 1
    wb.close()
    return headers or [], n, (max(states, key=states.get) if states else "")


def survey(xlsx, learn):
    """Return [(tab, rows, state, schema)] for tabs holding addresses. Learns an
    unknown layout via resolve_schema.py when a key is available."""
    if is_flat(xlsx):
        # one table, so one market, named from the file
        names = [os.path.splitext(os.path.basename(xlsx))[0]]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        names = [ws.title for ws in wb.worksheets]
        wb.close()

    out = []
    for tab in names:
        if tab.strip().lower() in SKIP_TABS:
            continue
        headers, n, state = read_headers(xlsx, tab)
        if not headers or n == 0:
            continue
        schema, m = build_mapping(headers)
        if not (m.get("street") or m.get("_f37")):
            if not learn:
                print(f"  ! '{tab}': layout not recognized, skipped "
                      f"(no ANTHROPIC_API_KEY, so it could not be learned)")
                continue
            print(f"  '{tab}': layout not recognized -- resolving", flush=True)
            r = sh([sys.executable, os.path.join(HERE, "resolve_schema.py"),
                    xlsx, "--tab", tab], check=False)
            if r.returncode != 0:
                print(f"  ! '{tab}': could not resolve the layout, skipped")
                continue
            headers, n, state = read_headers(xlsx, tab)
            schema, m = build_mapping(headers)
            if not (m.get("street") or m.get("_f37")):
                print(f"  ! '{tab}': still unreadable after resolving, skipped")
                continue
        out.append((tab, n, state, schema))
    return out


def explain_unreadable(path, err):
    """Turn a library exception into something an uploader can act on."""
    import zipfile
    name = os.path.basename(path)
    if isinstance(err, zipfile.BadZipFile):
        return (f"{name} is not a real spreadsheet file. The name ends in "
                f".xlsx but the contents are something else -- most often a "
                f"Google Sheet, or a PDF that got renamed. Open it, then "
                f"File > Download > Microsoft Excel (.xlsx), and upload that.")
    if isinstance(err, KeyError):
        return f"{name}: expected sheet is missing ({err})"
    return f"{name}: {type(err).__name__}: {err}"


def qa_summary(final_csv):
    import csv
    counts, zips = {}, set()
    for r in csv.DictReader(open(final_csv)):
        counts[r["qa_flag"]] = counts.get(r["qa_flag"], 0) + 1
        if r["lat"]:
            zips.add(r["zip"])
    held = counts.get("NO_MATCH", 0) + counts.get("FAR_FROM_ZIP_SEVERE", 0)
    return {"rows": sum(counts.values()), "held": held, "zips": len(zips),
            "mapped": sum(counts.values()) - held}


def build_market(xlsx, tab, state, work, site, with_csv, product=""):
    # URL is <product>/<market>, so two products can carry the same DMA tab and
    # stay separate maps. The work directory flattens that with "__", because a
    # slash there would create nesting the pruning logic would have to walk.
    tab_slug = slugify(tab)
    slug = f"{product}/{tab_slug}" if product else tab_slug
    wname = f"{product}__{tab_slug}" if product else tab_slug
    name = wname.replace("-", "_").replace("__", "_X_").upper()
    dest, wdir = os.path.join(site, *slug.split("/")), os.path.join(work, wname)
    os.makedirs(wdir, exist_ok=True)
    print(f"  {tab} -> /{slug}/", end="", flush=True)

    def p(script, *args):
        if VERBOSE:
            print()
        return sh([sys.executable, os.path.join(HERE, script), *args],
                  cwd=wdir, quiet=not VERBOSE)

    p("clean_addresses.py", xlsx, "--name", name, "--sheet", tab, "--state", state)

    geo_in = os.path.join(wdir, f"{name}_geocode_input.csv")
    geocoded = None
    if os.path.exists(geo_in) and os.path.getsize(geo_in) > 0:
        print(f"  geocoding {sum(1 for _ in open(geo_in)):,}", end="", flush=True)
        p("censusgeocode_full.py", geo_in, "--out",
          os.path.join(wdir, f"{name}_geocoded.csv"))
        geocoded = os.path.join(wdir, f"{name}_geocoded.csv")

    fin = ["finalize_addresses.py", os.path.join(wdir, f"{name}_addresses_clean.csv"),
           "--name", name]
    if geocoded:
        fin += ["--geocoded", geocoded]
    p(*fin)

    final_csv = os.path.join(wdir, f"{name}_addresses_final.csv")
    p("build_address_map.py", final_csv, "--name", name,
      "--title", titleize(tab, product))

    os.makedirs(dest, exist_ok=True)
    shutil.copy(os.path.join(wdir, f"{name}_address_map.html"),
                os.path.join(dest, "index.html"))
    # the audit CSV stays local -- internal record IDs and QA columns don't need
    # to sit on a public URL
    if with_csv:
        shutil.copy(final_csv, os.path.join(dest, f"{tab_slug}-addresses.csv"))
    print("  done")
    return slug, wname, qa_summary(final_csv)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("inputs", nargs="*", default=None,
                    help="spreadsheets, or a folder of them (default: inbox/)")
    ap.add_argument("--tabs", nargs="*", default=None)
    ap.add_argument("--site", default="site")
    ap.add_argument("--work", default="work")
    ap.add_argument("--with-csv", action="store_true")
    ap.add_argument("-v", "--verbose", action="store_true")
    a = ap.parse_args()

    global VERBOSE
    VERBOSE = a.verbose

    found = discover(a.inputs or ["inbox"])
    if not found:
        raise SystemExit("no spreadsheets found -- put one in inbox/<product>/")

    site, work = os.path.abspath(a.site), os.path.abspath(a.work)
    os.makedirs(site, exist_ok=True)
    os.makedirs(work, exist_ok=True)
    learn = bool(os.environ.get("ANTHROPIC_API_KEY", "").strip())

    built, failed, unreadable, seen, sources = [], [], [], {}, {}
    for product, path in found:
        name = os.path.basename(path)
        print(f"\nreading {product}/{name}")
        try:
            tabs = survey(os.path.abspath(path), learn)
        except Exception as e:
            why = explain_unreadable(path, e)
            print(f"  ! cannot read this file, skipped -- {why}")
            unreadable.append({"file": name, "reason": why})
            continue
        if a.tabs:
            want = {t.strip().lower() for t in a.tabs}
            tabs = [t for t in tabs if t[0].strip().lower() in want]
        if not tabs:
            print("  no address tabs found")
            continue
        print("  " + ", ".join(f"{t}({n:,}, {s})" for t, n, _, s in tabs))
        for tab, _, state, _ in tabs:
            key = f"{product}/{slugify(tab)}"
            if key in seen:
                print(f"  {tab}: also in {seen[key]} -- this file wins")
            seen[key] = name
            try:
                slug, wname, qa = build_market(os.path.abspath(path), tab, state,
                                               work, site, a.with_csv, product)
                built.append((tab, slug, qa))
                sources[wname] = name
            except SystemExit as e:
                # one unusable market must not take the others down with it
                print(f"\n  ! {tab}: build failed, skipped -- {e}")
                failed.append((tab, str(e)))

    js = __import__("json")
    with open(os.path.join(work, "_sources.json"), "w") as f:
        js.dump(sources, f, indent=2)
    # work/ is cached between CI runs so geocode checkpoints survive, which means
    # it holds a directory for every market ever built. Record this run's markets
    # so validate.py reports on what was actually published rather than on
    # everything the cache happens to be carrying.
    with open(os.path.join(work, "_built.json"), "w") as f:
        js.dump(sorted(sources.keys()), f, indent=2)

    # prune leftovers, so the cache cannot grow without limit. Skipped when only
    # some tabs were requested, since then the others are absent on purpose.
    if not a.tabs:
        live = set(sources)
        for entry in sorted(os.listdir(work)):
            d = os.path.join(work, entry)
            if not os.path.isdir(d) or entry in live:
                continue
            shutil.rmtree(d, ignore_errors=True)
            print(f"  pruned stale work dir: {entry}")
    if unreadable:
        with open(os.path.join(work, "_unreadable.json"), "w") as f:
            __import__("json").dump(unreadable, f, indent=2)
    if failed or unreadable:
        print("\n" + "-" * 62)
        for tab, why in failed:
            print(f"FAILED  {tab}: {why}")
        for u in unreadable:
            print(f"SKIPPED {u['file']}")
    if not built:
        raise SystemExit("nothing built")
    print("\n" + "-" * 62)
    for tab, slug, qa in built:
        print(f"{tab:<28} {qa['mapped']:>7,} mapped  {qa['zips']:>3} ZIPs  "
              f"{qa['held']:>5,} held out   /{slug}/")
    print(f"\nsite/  {len(built)} market(s) built"
          + (f", {len(failed)} failed" if failed else "") + "   audit CSVs in work/")


if __name__ == "__main__":
    main()
