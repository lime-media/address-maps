#!/usr/bin/env python3
"""
clean_addresses.py -- Stage 1 of the address-map pipeline.

Reads an address spreadsheet (.xlsx / .xls / .csv), auto-detects the column
schema, standardizes the address text, and emits two files:

    <NAME>_addresses_clean.csv   full audit table (every input row, one line each)
    <NAME>_geocode_input.csv     Census-ready, NO header, ONLY rows missing coords
                                 -> id, street, city, state, zip

Usage:
    python3 clean_addresses.py INPUT --name DEMO [--state TX] [--sheet all|NAME]

Schemas recognized by header (case/space-insensitive):
    Raw AT&T   SERVICE_ADDRESS · SERVICE_CITY · SERVICE_STATE · SERVICE_POSTAL_CD
               · SERVICE_APT_NUM (present = multi-family) · GLID
    Alloy RG   T_ADDRESS · T_CITY · T_STATE · T_ZIP CODE · T_LAT · T_LON · T_UNIQUE ID
    F37        F37 ID · Street Number/Name/Type · Sub Unit Type/Number · Postal Code
    Generic    anything containing address/street, city, state, zip/postal, lat, lon

Nothing is dropped here. Units are captured in their own column and kept OUT of
the `street` field sent to the geocoder -- units hurt Census match rate. Rows that
already carry coordinates are passed through and never geocoded.
"""
import argparse
import json
import csv
import os
import re
import sys

# ---------------------------------------------------------------- USPS text rules
SUFFIX = {
    "STREET": "ST", "STR": "ST", "ST.": "ST",
    "AVENUE": "AVE", "AV": "AVE", "AVEN": "AVE",
    "BOULEVARD": "BLVD", "BLVD.": "BLVD", "BOUL": "BLVD",
    "ROAD": "RD", "DRIVE": "DR", "DRV": "DR",
    "LANE": "LN", "COURT": "CT", "CIRCLE": "CIR", "CIRC": "CIR",
    "PLACE": "PL", "TERRACE": "TER", "TERR": "TER",
    "PARKWAY": "PKWY", "PKY": "PKWY", "PARKWY": "PKWY",
    "HIGHWAY": "HWY", "HIWAY": "HWY",
    "TRAIL": "TRL", "TRAILS": "TRL",
    "SQUARE": "SQ", "LOOP": "LOOP", "PASS": "PASS", "PATH": "PATH",
    "CROSSING": "XING", "COVE": "CV", "BEND": "BND", "RIDGE": "RDG",
    "POINT": "PT", "RUN": "RUN", "WAY": "WAY", "WY": "WAY",
    "EXPRESSWAY": "EXPY", "FREEWAY": "FWY", "TURNPIKE": "TPKE",
    "PLAZA": "PLZ", "GARDENS": "GDNS", "GARDEN": "GDN",
    "HEIGHTS": "HTS", "HOLLOW": "HOLW", "MEADOWS": "MDWS", "MEADOW": "MDW",
    "CREEK": "CRK", "SPRINGS": "SPGS", "SPRING": "SPG",
    "VALLEY": "VLY", "VIEW": "VW", "VILLAGE": "VLG", "JUNCTION": "JCT",
    "EXTENSION": "EXT", "ISLAND": "IS", "LAKE": "LK", "MOUNT": "MT",
    "MOUNTAIN": "MTN", "RIVER": "RIV", "SHORE": "SHR", "STATION": "STA",
    "ESTATES": "ESTS", "LANDING": "LNDG", "FOREST": "FRST", "GLEN": "GLN",
}
DIRECTIONAL = {
    "NORTH": "N", "SOUTH": "S", "EAST": "E", "WEST": "W",
    "NORTHEAST": "NE", "NORTHWEST": "NW", "SOUTHEAST": "SE", "SOUTHWEST": "SW",
    "NO": "N", "SO": "S",
}
UNIT_WORDS = re.compile(
    r"\b(APT|APARTMENT|UNIT|STE|SUITE|BLDG|BUILDING|TRLR|TRAILER|LOT|SPC|SPACE|"
    r"RM|ROOM|FL|FLOOR|#)\b\.?\s*([A-Z0-9\-]+)?", re.I)


def standardize_street(raw):
    """Uppercase, strip punctuation/noise, abbreviate directionals + suffixes.

    Returns (street_without_unit, unit_text).
    """
    if not raw:
        return "", ""
    s = str(raw).upper().strip()
    s = s.replace("&", " AND ")
    s = re.sub(r"[.,;:\"']", " ", s)

    # pull a trailing/embedded unit designator out of the street line
    unit = ""
    m = UNIT_WORDS.search(s)
    if m:
        unit = (m.group(2) or m.group(1)).strip()
        s = (s[:m.start()] + " " + s[m.end():])
    s = re.sub(r"\s+", " ", s).strip()

    toks = s.split(" ")
    out = []
    for i, t in enumerate(toks):
        # directionals only at the head or tail -- "NORTH BEND RD" keeps its middle word
        if t in DIRECTIONAL and (i == 0 or i == len(toks) - 1 or i == 1):
            out.append(DIRECTIONAL[t])
        elif t in SUFFIX:
            out.append(SUFFIX[t])
        else:
            out.append(t)
    return " ".join(out).strip(), unit.strip()


def clean_zip(v):
    if v is None:
        return ""
    z = re.sub(r"\D", "", str(v))
    if len(z) > 5:
        z = z[:5]
    return z.zfill(5) if z else ""


def to_float(v):
    try:
        f = float(str(v).strip())
        return f if f != 0 else None
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------- schema detection
def norm(h):
    return re.sub(r"[^A-Z0-9]", "", str(h or "").upper())


def find(headers, *candidates, contains=False):
    """Return the ORIGINAL header matching any candidate, else None."""
    nmap = {norm(h): h for h in headers}
    for c in candidates:
        if norm(c) in nmap:
            return nmap[norm(c)]
    if contains:
        for c in candidates:
            for n, orig in nmap.items():
                if norm(c) in n:
                    return orig
    return None


def load_learned(path=None):
    """Layouts learned from files we've seen before, stored as data.

    schemas.json sits beside the pipeline and grows over time. A layout is matched
    when every header in its fingerprint is present, and the most specific
    (longest) fingerprint wins. This is checked AFTER the built-in schemas below
    and BEFORE the generic fallback, so hand-verified layouts always take
    precedence and learned ones only fill genuine gaps.
    """
    path = path or os.environ.get("ADDRESS_MAP_SCHEMAS") or \
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "schemas.json")
    if not os.path.exists(path):
        return []
    try:
        with open(path) as f:
            return json.load(f).get("layouts", [])
    except (ValueError, OSError) as e:
        print(f"  ! could not read {path}: {e}", file=sys.stderr)
        return []


def match_learned(headers, layouts):
    """Return (name, mapping) for the most specific matching layout, else None."""
    present = {norm(h) for h in headers}
    best = None
    for lay in layouts:
        fp = [norm(x) for x in lay.get("fingerprint", [])]
        if not fp or not all(x in present for x in fp):
            continue
        if best is None or len(fp) > len(best[0]):
            best = (fp, lay)
    if not best:
        return None
    lay = best[1]
    # a learned map records ORIGINAL header names; keep only ones actually here
    m = {k: (v if v in headers else None) for k, v in lay.get("map", {}).items()}
    if not m.get("street"):
        return None
    return f"learned: {lay.get('name', 'unnamed')}", m


def build_mapping(headers):
    """Return (schema_name, mapping dict). mapping values are header names or None.

    F37 uses composite street parts, flagged by mapping['_f37'] = [parts].
    """
    H = headers
    if find(H, "SERVICE_POSTAL_CD"):
        return "Raw AT&T", {
            "id": find(H, "GLID", "SERVICE_ID"),
            "street": find(H, "SERVICE_ADDRESS", "SERVICE_STREET_ADDRESS", "SERVICE_ADDR"),
            "unit": find(H, "SERVICE_APT_NUM"),
            "city": find(H, "SERVICE_CITY"),
            "state": find(H, "SERVICE_STATE", "SERVICE_STATE_CD"),
            "zip": find(H, "SERVICE_POSTAL_CD"),
            "lat": find(H, "LATITUDE", "LAT"),
            "lon": find(H, "LONGITUDE", "LON", "LNG"),
        }
    if find(H, "T_ZIP CODE", "T_ZIPCODE"):
        return "Alloy RG", {
            "id": find(H, "T_UNIQUE ID", "T_UNIQUEID"),
            "street": find(H, "T_ADDRESS", "T_STREET", "T_ADDRESS_1"),
            "unit": None,
            "city": find(H, "T_CITY"),
            "state": find(H, "T_STATE"),
            "zip": find(H, "T_ZIP CODE", "T_ZIPCODE"),
            "lat": find(H, "T_LAT", "T_LATITUDE"),
            "lon": find(H, "T_LON", "T_LONGITUDE", "T_LNG"),
        }
    if find(H, "F37 ID", "F37ID"):
        parts = [find(H, "Street Number"), find(H, "Pre Directional", "Street Pre Directional"),
                 find(H, "Street Name"), find(H, "Street Type", "Street Suffix"),
                 find(H, "Post Directional", "Street Post Directional")]
        return "F37", {
            "id": find(H, "F37 ID", "F37ID"),
            "_f37": [p for p in parts if p],
            "street": None,
            "unit": find(H, "Sub Unit Number", "Sub Unit Type"),
            "city": find(H, "City", "Municipality"),
            "state": find(H, "State", "State Code"),
            "zip": find(H, "Postal Code", "ZIP"),
            "lat": find(H, "Latitude", "Lat"),
            "lon": find(H, "Longitude", "Lon"),
        }
    learned = match_learned(H, load_learned())
    if learned:
        return learned
    # generic
    return "Generic", {
        "id": find(H, "id", "uniqueid", "recordid", "rowid", contains=True),
        "street": find(H, "address1", "streetaddress", "address", "street", "addr", contains=True),
        "unit": find(H, "unit", "apt", "suite", "secondary", contains=True),
        "city": find(H, "city", "town", "municipality", contains=True),
        "state": find(H, "state", "st", "province", contains=True),
        "zip": find(H, "zip", "zipcode", "postalcode", "postal", contains=True),
        "lat": find(H, "latitude", "lat", contains=True),
        "lon": find(H, "longitude", "long", "lon", "lng", contains=True),
    }


# ---------------------------------------------------------------- readers
def read_rows(path, sheet="all"):
    """Yield (sheet_name, headers, list-of-dict rows) for each sheet/table."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".txt", ".tsv"):
        delim = "\t" if ext == ".tsv" else ","
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            rd = csv.DictReader(f, delimiter=delim)
            yield os.path.splitext(os.path.basename(path))[0], rd.fieldnames or [], list(rd)
        return

    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    names = wb.sheetnames if sheet == "all" else [sheet]
    for name in names:
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        headers = None
        for row in it:                       # tolerate a blank/banner row above the header
            if row and sum(1 for c in row if c not in (None, "")) >= 2:
                headers = [str(c).strip() if c is not None else "" for c in row]
                break
        if not headers:
            continue
        rows = []
        for row in it:
            if row is None or all(c in (None, "") for c in row):
                continue
            rows.append({h: v for h, v in zip(headers, row)})
        yield name, headers, rows
    wb.close()


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("--name", required=True, help="output prefix, e.g. DALLAS_JUL26")
    ap.add_argument("--state", default="", help="2-letter fallback if the file has no state column")
    ap.add_argument("--sheet", default="all", help="'all' (default) or a single sheet name")
    ap.add_argument("--outdir", default=".")
    a = ap.parse_args()

    clean_path = os.path.join(a.outdir, f"{a.name}_addresses_clean.csv")
    geo_path = os.path.join(a.outdir, f"{a.name}_geocode_input.csv")

    total = kept = have_coords = need_geo = no_street = mfh = 0
    seq = 0
    seen_ids = set()

    with open(clean_path, "w", newline="") as cf, open(geo_path, "w", newline="") as gf:
        cw = csv.writer(cf)
        cw.writerow(["id", "source_group", "street", "unit", "city", "state",
                     "zip", "lat", "lon", "coord_source", "mfh"])
        gw = csv.writer(gf)

        for sheet_name, headers, rows in read_rows(a.input, a.sheet):
            schema, m = build_mapping(headers)
            if not (m.get("street") or m.get("_f37")):
                print(f"  ! sheet '{sheet_name}': no address column found -- skipped "
                      f"(headers: {', '.join(str(h) for h in headers[:8])})", file=sys.stderr)
                continue
            print(f"  sheet '{sheet_name}': {len(rows):,} rows | schema = {schema}")

            for r in rows:
                total += 1
                if m.get("_f37"):
                    raw_street = " ".join(str(r.get(p) or "").strip() for p in m["_f37"]).strip()
                else:
                    raw_street = r.get(m["street"])
                street, unit_from_text = standardize_street(raw_street)
                if not street:
                    no_street += 1
                    continue

                unit = str(r.get(m["unit"]) or "").strip().upper() if m.get("unit") else ""
                unit = unit or unit_from_text
                is_mfh = bool(unit)
                if is_mfh:
                    mfh += 1

                city = re.sub(r"\s+", " ", str(r.get(m["city"]) or "").strip()).title() if m.get("city") else ""
                state = (str(r.get(m["state"]) or "").strip().upper()[:2] if m.get("state") else "") or a.state.upper()
                zipc = clean_zip(r.get(m["zip"])) if m.get("zip") else ""

                lat = to_float(r.get(m["lat"])) if m.get("lat") else None
                lon = to_float(r.get(m["lon"])) if m.get("lon") else None
                if lat is not None and not (-90 <= lat <= 90):
                    lat = None
                if lon is not None and not (-180 <= lon <= 180):
                    lon = None
                # a lat/lon swap shows up as an out-of-CONUS pair -- fix the obvious case
                if lat is not None and lon is not None and lon > 0 and lat < 0:
                    lat, lon = lon, lat

                rid = str(r.get(m["id"]) or "").strip() if m.get("id") else ""
                if not rid or rid in seen_ids:
                    seq += 1
                    rid = f"R{seq:08d}"
                seen_ids.add(rid)

                if lat is not None and lon is not None:
                    have_coords += 1
                    cw.writerow([rid, sheet_name, street, unit, city, state, zipc,
                                 f"{lat:.6f}", f"{lon:.6f}", "file", is_mfh])
                else:
                    need_geo += 1
                    cw.writerow([rid, sheet_name, street, unit, city, state, zipc,
                                 "", "", "pending", is_mfh])
                    gw.writerow([rid, street, city, state, zipc])   # street WITHOUT unit
                kept += 1

    print(f"\n{total:,} input rows -> {kept:,} usable "
          f"({no_street:,} dropped for no street address)")
    print(f"  {have_coords:,} already carry coordinates")
    print(f"  {need_geo:,} need geocoding")
    print(f"  {mfh:,} carry a unit/apt (multi-family signal)")
    print(f"\nwrote {clean_path}")
    print(f"wrote {geo_path}")
    if need_geo:
        print(f"\nNEXT: python3 censusgeocode_full.py {geo_path}")
    else:
        print(f"\nNEXT (no geocoding needed): python3 finalize_addresses.py "
              f"{clean_path} --name {a.name}")


if __name__ == "__main__":
    main()
